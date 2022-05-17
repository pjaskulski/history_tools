""" tworzenie list regestów ze Słownika Historyczno-Geograficznego """

import re
from pathlib import Path
import openpyxl
from tools import create_sheet, get_regesty, get_year_from_regest
from tools import get_source_from_regest, get_contents_from_regest


slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw'
}


def process_point(sheet, m_id, m_nazwa, point_num, point_text):
    """przetwarzanie treści punktu z zapisem do wskazanego arkusza
       sheet - wskazanie na arkusz
       m_id - identyfikator miejscowości
       nazwa - nazwa miejscowości
       point_num - numer punktu
       point_text - treść punktu
    """
    global regesty_count
    global place_owner

    reg_list = []
    point = point_num

    # czyszczenie z pozostałości HTML-a
    pattern = r'<.*?>'
    point_text = re.sub(pattern, '', point_text)
    point_text = point_text.replace(r'\n', ' ')

    reg_list, ownership = get_regesty(point_text)
    # własność
    if point_num == 3:
        place_owner[m_nazwa] = ownership

    # specjalna obsługa dla p. 1
    if point_num == 1:
        new_reg_list = []
        for item in reg_list:
            if ")," in item:
                tmp = item.split("),")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            elif ")." in item:
                tmp = item.split(").")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            else:
                new_reg_list.append(item)
        reg_list = new_reg_list

    for item in reg_list:
        # if 'n. par. Nowa Słupia (DLb. II 490) [' in item:
        #     print()
        item = item.strip()
        year = get_year_from_regest(item)
        source = get_source_from_regest(item)
        contents = get_contents_from_regest(item, year, source, m_nazwa)
        item = item.replace('[@', '[').replace('@]',']')
        tmp_regest = [m_id, m_nazwa, point, year, contents, source, item]
        regest = tuple(tmp_regest)
        sheet.append(regest)
        regesty_count[point_num] += 1


# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':
    regesty_count = {1:0, 2:0, 3:0, 4:0, 5:0, 6:0}

    for nr in range(1, 13):
        print(f"Słownik: {slowniki[nr]} ({nr})")

        place_owner = {}

        output_path = Path('.').parent / f'output/wyniki/regesty_{slowniki[nr]}.xlsx'
        input_path = Path('.').parent / f'output/wyniki/hasla_{nr}_pkt.xlsx'

        wb_hasla = openpyxl.load_workbook(input_path)
        ws_hasla = wb_hasla.active

        wb = openpyxl.Workbook()
        ws = wb.active

        point_sheet_columns = ['Miejscowość_Id', 'Nazwa', 'Punkt', 'Rok',
                               'Treść', 'Źródła', 'Cały Regest']
        regest_sheet_1 = create_sheet(wb=wb, title="Regesty p. 1", columns=point_sheet_columns)
        regest_sheet_2 = create_sheet(wb=wb, title="Regesty p. 2", columns=point_sheet_columns)
        regest_sheet_3 = create_sheet(wb=wb, title="Regesty p. 3", columns=point_sheet_columns)
        regest_sheet_4 = create_sheet(wb=wb, title="Regesty p. 4", columns=point_sheet_columns)
        regest_sheet_5 = create_sheet(wb=wb, title="Regesty p. 5", columns=point_sheet_columns)
        regest_sheet_6 = create_sheet(wb=wb, title="Regesty p. 6", columns=point_sheet_columns)

        # kolumny w źródłowym pierwszym arkuszu
        col_names = {}
        nr_col = 0
        for column in ws_hasla.iter_cols(1, ws_hasla.max_column):
            col_names[column[0].value] = nr_col
            nr_col += 1

        # przetwarzanie wierszy
        max_hasla = ws_hasla.max_row
        licznik = 0
        for row in ws_hasla.iter_rows(2, max_hasla):
            licznik += 1
            print(f'Przetwarzam wiersz {licznik} z {max_hasla}.')
            miejsce_id = row[col_names['id']].value
            nazwa = row[col_names['header']].value
            nazwa = re.sub(r'<.*?>', '', nazwa)
            p1 = row[col_names[1]].value
            p2 = row[col_names[2]].value
            p3 = row[col_names[3]].value
            p4 = row[col_names[4]].value
            p5 = row[col_names[5]].value
            p6 = row[col_names[6]].value

            # Punkt 1
            if p1 is not None:
                process_point(regest_sheet_1, miejsce_id, nazwa, 1, p1)
            if p2 is not None:
                process_point(regest_sheet_2, miejsce_id, nazwa, 2, p2)
            if p3 is not None:
                process_point(regest_sheet_3, miejsce_id, nazwa, 3, p3)
            if p4 is not None:
                process_point(regest_sheet_4, miejsce_id, nazwa, 4, p4)
            if p5 is not None:
                process_point(regest_sheet_5, miejsce_id, nazwa, 5, p5)
            if p6 is not None:
                process_point(regest_sheet_6, miejsce_id, nazwa, 6, p6)

        # usunięcie pustego pierwszego arkusza
        sheet1 = wb['Sheet']
        wb.remove(sheet1)

        # arkusz z własnością
        wlasnosc_sheet = create_sheet(wb=wb, title="Własność", columns=['Miejscowość', 'Własność'])
        for key, value in place_owner.items():
            wiersz = (key, value)
            wlasnosc_sheet.append(wiersz)

        # zapis pliku xlsx
        wb.save(filename=output_path)

    suma = 0
    for i in range(1, 7):
        print(f"p.{i}", regesty_count[i])
        suma += regesty_count[i]
    print(f"razem: {suma}")
        